#!/usr/bin/env python3
"""Import supplier price lists into local Karzar catalog (SQL).

Sources under DOCS/Price:
  - Asal Excel (INSIZE) — TOMAN column as-is
  - Brand PDFs — prices labeled Rial → store as Toman (÷10)

Usage:
  .venv312/bin/python scripts/import_price_lists.py --dry-run
  .venv312/bin/python scripts/import_price_lists.py --apply
  .venv312/bin/python scripts/import_price_lists.py --apply --phase asal
  .venv312/bin/python scripts/import_price_lists.py --apply --phase pdf
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path

import psycopg2
from psycopg2.extras import RealDictCursor

PRICE_DIR = Path("/home/moahmmad/Projects/Karzar/DOCS/Price")
REPORT_DIR = Path(__file__).resolve().parents[1] / "data"
INSIZE_BRAND_ID = 3

# Latest file per family (plan).
ASAL_XLSX = PRICE_DIR / "آصال تومان 27-04-1405.xlsx"

# (brand_hint, brand_id|None, path) — brand_id scopes matching to avoid cross-brand SKU collisions.
PDF_FILES: list[tuple[str, int | None, Path]] = [
    ("Dasqua", 4, PRICE_DIR / "لیست قیمت داسکوا 14050212.pdf"),
    ("Chumpower", 33, PRICE_DIR / "لیست چام پاور 14050215.pdf"),
    ("ASTPOWER", 13, PRICE_DIR / "سه نظام ای اس تی 14050216.pdf"),
    ("ASTPOWER", 13, PRICE_DIR / "قلاویزگیر AST 14050401.pdf"),
    ("ASTPOWER", 13, PRICE_DIR / "گیره مارک AST POWER 14050222.pdf"),
    ("ASTPOWER", 13, PRICE_DIR / "لیست قیمت پول استادگیر AST POWER 14050219.pdf"),
    ("ASTPOWER", 13, PRICE_DIR / "لیست قیمت دریل مگنت 14050131.pdf"),
    ("ASTPOWER", 13, PRICE_DIR / "لیست قیمت کردریل 14050131.pdf"),
    ("ASTPOWER", 13, PRICE_DIR / "لیست قیمت گردبر دریل مگنت 14050206.pdf"),
    ("ASTPOWER", 13, PRICE_DIR / "قلاویز زن اتومات 14050222.pdf"),
    ("ASTPOWER", 13, PRICE_DIR / "قلاویززن برقی 14050214.pdf"),
    ("ASTPOWER", 13, PRICE_DIR / "لیست قلاویززن بادی 14050210.pdf"),
    ("ET", 13, PRICE_DIR / "قلاویز ET 14050216.pdf"),  # ET taps sold under ASTPOWER catalog
    ("UTEX", 24, PRICE_DIR / "مته یوتکس  14050209.pdf"),  # diameter matrix — usually no SKU match
    ("drill", None, PRICE_DIR / "مته ته کونیک 14050209.pdf"),
    ("drill", None, PRICE_DIR / "لیست قیمت  مته برگی 14050203.pdf"),
    ("drill", None, PRICE_DIR / "لیست مته مرغک و مته خزینه 14050331.pdf"),
    ("misc", 13, PRICE_DIR / "تایکوپ و صفحه گردان 14050409.pdf"),
    ("misc", 13, PRICE_DIR / "کلاهک و دنباله 14050209.pdf"),
    ("misc", None, PRICE_DIR / "لیست قیمت روغن اب صابون 14050222.pdf"),
    ("misc", None, PRICE_DIR / "لیست قیمت  فرز انگشتی 14050216.pdf"),
    ("misc", 13, PRICE_DIR / "لیست قیمت ابزار تیز کن 14050214.pdf"),
    ("3Keego", 35, PRICE_DIR / "گردبر کیگو 14050214.pdf"),
]

# Minimum valid unit price in Rial (PDF). Below this → reject (row index / fragment).
MIN_RIAL_PRICE = 100_000
# Minimum stored Toman after conversion.
MIN_TOMAN_PRICE = 10_000

# SKU-like tokens and comma-grouped prices (rial).
SKU_TOKEN_RE = re.compile(
    r"(?i)(?<![A-Z0-9/])("
    r"[A-Z]{1,6}-?[A-Z0-9]*/?[A-Z0-9][A-Z0-9._×x*]{1,40}"
    r"|\d{3,5}-\d{2,5}[A-Z]{0,3}"
    r"|\d{4,8}(?:/\d{1,3})?"
    r")(?![A-Z0-9-])"
)
PRICE_RE = re.compile(r"(?<![\d.])(\d{1,3}(?:,\d{3}){1,4}|\d{4,12})(?![\d.])")
JUNK_SKU = re.compile(
    r"(?i)^(mm|cm|kg|din|hsse|hss|iso|unf|unc|nc|nf|b\d{1,2}|m\d+(\.\d+)?|"
    r"ip\d+|vh\d+|hv\d+|tc\d+|page|row)$"
)


@dataclass
class PriceRow:
    code: str
    price_toman: int | None  # None => inquiry (0 rial)
    source: str
    raw_price: str = ""


def db_connect():
    return psycopg2.connect(
        host=os.getenv("POSTGRES_SERVER", "127.0.0.1"),
        port=int(os.getenv("POSTGRES_PORT", "5435")),
        user=os.getenv("POSTGRES_USER", "postgres"),
        password=os.getenv("POSTGRES_PASSWORD", "your_secure_password_here"),
        dbname=os.getenv("POSTGRES_DB", "karzar_db"),
    )


def norm(code: str) -> str:
    c = str(code or "").strip().upper()
    c = c.replace(" ", "").replace("_", "-").replace("×", "X").replace("*", "X")
    c = c.replace("‐", "-").replace("–", "-").replace("—", "-")
    return c


def code_variants(code: str) -> list[str]:
    c = norm(code)
    out = [c]
    m = re.match(r"^(\d{3,5}-\d+)([A-Z]+)$", c)
    if m:
        out.append(m.group(1))
    m2 = re.match(r"^(\d{3,5}-\d+[0-9])([A-Z]+)$", c)
    if m2 and m2.group(1) not in out:
        out.append(m2.group(1))
    # strip trailing single letter after digits already handled; also AST- prefixes
    if c.startswith("AST-"):
        out.append(c[4:])
    return list(dict.fromkeys(out))


def parse_price_number(raw: str) -> int | None:
    s = str(raw).strip().replace(",", "").replace("٬", "")
    if not s:
        return None
    try:
        val = int(float(s))
    except (TypeError, ValueError):
        return None
    return val


def load_asal_excel(path: Path) -> list[PriceRow]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[PriceRow] = []
    for r in range(2, (ws.max_row or 0) + 1):
        code = ws.cell(r, 1).value
        toman = ws.cell(r, 4).value
        if code is None:
            continue
        key = norm(code)
        if not key:
            continue
        if toman is None:
            continue
        price = parse_price_number(toman)
        if price is None:
            continue
        if price <= 0:
            rows.append(PriceRow(code=key, price_toman=None, source=path.name, raw_price=str(toman)))
        else:
            rows.append(PriceRow(code=key, price_toman=price, source=path.name, raw_price=str(toman)))
    # last wins on duplicate codes
    by: dict[str, PriceRow] = {}
    for row in rows:
        by[row.code] = row
    return list(by.values())


def pdftotext(path: Path) -> str:
    proc = subprocess.run(
        ["pdftotext", "-layout", str(path), "-"],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return proc.stdout


def _looks_like_sku(token: str) -> bool:
    t = norm(token)
    if len(t) < 3 or len(t) > 48:
        return False
    if JUNK_SKU.match(t):
        return False
    if re.fullmatch(r"\d{1,3}", t):
        return False
    # size ranges mistaken as codes
    if re.search(r"(?i)\d+-\d+\s*MM$", t) or re.search(r"(?i)^\d+-\d+MM$", t):
        return False
    if re.fullmatch(r"\d+-\d+MM", t, re.I):
        return False
    # pure decimals / sizes without letters
    if re.fullmatch(r"\d+[./]\d+", t) and "-" not in t and not re.search(r"[A-Z]", t):
        return False
    if not re.search(r"\d", t):
        return False
    # bare short numeric codes are ambiguous fragments (e.g. 4410 from 4410-1105-A)
    if re.fullmatch(r"\d{3,5}", t):
        return False
    return True


def parse_pdf_prices(path: Path, brand_hint: str) -> list[PriceRow]:
    """Extract (sku, rial_price) pairs from layout text; convert to toman."""
    text = pdftotext(path)
    rows: list[PriceRow] = []
    lines = text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        if i + 1 < len(lines):
            a = line.rstrip()
            b = lines[i + 1].lstrip()
            if re.search(r"[-/A-Z0-9]$", a, re.I) and re.match(r"^[A-Z0-9]", b, re.I):
                if not re.match(r"^\d+\s", b):
                    line = a + b
                    i += 1

        prices = [(m.start(), m.group(1), parse_price_number(m.group(1))) for m in PRICE_RE.finditer(line)]
        money = []
        for start, raw, val in prices:
            if val is None:
                continue
            # Accept: comma-grouped amounts, explicit 0, or large plain integers
            if "," in raw or val == 0 or val >= MIN_RIAL_PRICE:
                money.append((start, raw, val))
        if not money:
            i += 1
            continue

        money.sort(key=lambda x: x[0])
        comma_prices = [m for m in money if "," in m[1]]
        zero_prices = [m for m in money if m[2] == 0]
        if comma_prices:
            candidates = comma_prices
        elif zero_prices and not any(m[2] and m[2] >= MIN_RIAL_PRICE for m in money):
            candidates = zero_prices
        else:
            candidates = [m for m in money if m[2] and m[2] >= MIN_RIAL_PRICE]
        if not candidates:
            i += 1
            continue

        _, raw_price, rial = candidates[-1]
        if rial is None:
            i += 1
            continue
        if rial != 0 and rial < MIN_RIAL_PRICE:
            i += 1
            continue

        left = line[: candidates[-1][0]]
        skus = [m.group(1) for m in SKU_TOKEN_RE.finditer(left) if _looks_like_sku(m.group(1))]
        if not skus:
            skus = [m.group(1) for m in SKU_TOKEN_RE.finditer(line) if _looks_like_sku(m.group(1))]
        if not skus:
            i += 1
            continue

        # Drop tokens that are strict prefixes of a longer token on the same line
        norms = [norm(s) for s in skus]
        filtered = []
        for s in norms:
            if any(o != s and o.startswith(s) and (o[len(s):len(s)+1] in "-/" or len(o) > len(s) + 1) for o in norms):
                continue
            filtered.append(s)
        if not filtered:
            filtered = norms

        def sku_score(t: str) -> tuple:
            return (
                1 if re.search(r"[A-Z]", t) and re.search(r"\d", t) else 0,
                1 if re.match(r"^\d{3,5}-\d+", t) else 0,
                1 if "-" in t or "/" in t else 0,
                len(t),
            )

        sku = max(filtered, key=sku_score)
        if rial == 0:
            toman = None
        else:
            toman = rial // 10
            if toman < MIN_TOMAN_PRICE:
                i += 1
                continue
        rows.append(
            PriceRow(
                code=sku,
                price_toman=toman,
                source=f"{brand_hint}:{path.name}",
                raw_price=raw_price,
            )
        )
        i += 1

    by: dict[str, PriceRow] = {}
    for row in rows:
        by[row.code] = row
    return list(by.values())


def load_products(conn) -> list[dict]:
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(
            """
            SELECT p.id, p.sku, p.name, p.brand_id, p.base_price, b.name AS brand_name
            FROM products p
            LEFT JOIN brands b ON b.id = p.brand_id
            WHERE p.deleted_at IS NULL
            """
        )
        return list(cur.fetchall())


def soft_key(code: str) -> str:
    """Normalize for fuzzy equality (slash/dash/space)."""
    return norm(code).replace("/", "-").replace(".", "")


def match_prices(
    products: list[dict],
    price_rows: list[PriceRow],
    *,
    brand_id: int | None = None,
) -> dict:
    excel_by = {r.code: r for r in price_rows}
    soft_index: dict[str, list[str]] = defaultdict(list)
    for code in excel_by:
        soft_index[soft_key(code)].append(code)
        for v in code_variants(code):
            soft_index[soft_key(v)].append(code)

    by_base: dict[str, list[str]] = defaultdict(list)
    for code in excel_by:
        for v in code_variants(code):
            by_base[v].append(code)

    matched = []
    unmatched_products = []
    methods: Counter = Counter()

    scoped = products
    if brand_id is not None:
        scoped = [p for p in products if p.get("brand_id") == brand_id]

    # Precompute name haystacks for name-contains matching
    name_map: list[tuple[dict, str]] = []
    for p in scoped:
        hay = soft_key(f"{p.get('sku') or ''} {p.get('name') or ''}")
        name_map.append((p, hay))

    used_codes: set[str] = set()
    matched_pids: set[int] = set()

    def try_bind(p, hit_code: str, method: str) -> bool:
        if p["id"] in matched_pids or hit_code not in excel_by:
            return False
        matched.append((p, excel_by[hit_code], method))
        methods[method] += 1
        used_codes.add(hit_code)
        matched_pids.add(p["id"])
        return True

    # Pass 1: SKU exact / variant
    for p in scoped:
        sku = norm(p.get("sku"))
        hit = None
        method = None
        if sku in excel_by:
            hit, method = sku, "exact"
        else:
            sk = soft_key(sku)
            soft_cands = list(dict.fromkeys(soft_index.get(sk) or []))
            if len(soft_cands) == 1:
                hit, method = soft_cands[0], "soft-exact"
            else:
                for v in code_variants(sku):
                    if v in excel_by:
                        hit, method = v, "sku-as-list-base"
                        break
                    cands = list(dict.fromkeys(by_base.get(v) or []))
                    if len(cands) == 1:
                        hit, method = cands[0], "unique-base"
                        break
                    if len(cands) > 1:
                        if sku in cands:
                            hit, method = sku, "exact-among-cands"
                        else:
                            letter = [c for c in cands if c.startswith(sku) and len(c) == len(sku) + 1]
                            if len(letter) == 1:
                                hit, method = letter[0], "sku+letter"
                            else:
                                method = "ambiguous"
                        break
        if hit and method != "ambiguous":
            try_bind(p, hit, method)
        elif method == "ambiguous":
            methods["ambiguous_skip"] += 1
            unmatched_products.append(p)
        else:
            unmatched_products.append(p)

    # Pass 2: price-list code appears uniquely in product name (brand-scoped)
    # Only for still-unmatched products / unused codes; require code length >= 5.
    still_unmatched = [p for p in unmatched_products if p["id"] not in matched_pids]
    unmatched_products = []
    unused_codes = [c for c in excel_by if c not in used_codes]

    for code in unused_codes:
        if len(code) < 5:
            continue
        needle = soft_key(code)
        if len(needle) < 5:
            continue
        hits = []
        for p, hay in name_map:
            if p["id"] in matched_pids:
                continue
            if needle in hay:
                hits.append(p)
        if len(hits) == 1:
            try_bind(hits[0], code, "name-contains")

    for p in still_unmatched:
        if p["id"] not in matched_pids:
            unmatched_products.append(p)

    unused_list = [excel_by[c] for c in excel_by if c not in used_codes]
    return {
        "matched": matched,
        "unmatched_products": unmatched_products,
        "unused_list_codes": unused_list,
        "methods": methods,
    }


def apply_updates(conn, matched, *, dry_run: bool) -> dict:
    to_update = []
    same = 0
    for p, row, method in matched:
        old = p.get("base_price")
        try:
            old_f = float(old) if old is not None else None
        except (TypeError, ValueError):
            old_f = None
        new = row.price_toman  # may be None
        if new is not None and new < MIN_TOMAN_PRICE:
            # Guard against fragment mis-parses
            continue
        if new is None and old_f is None:
            same += 1
            continue
        if new is not None and old_f is not None and abs(old_f - new) < 0.5:
            same += 1
            continue
        to_update.append((p, row, method, old))

    report = {
        "same": same,
        "to_update": len(to_update),
        "updated": [],
        "failed": [],
    }
    if dry_run:
        report["sample"] = [
            {
                "sku": p["sku"],
                "old": str(old) if old is not None else None,
                "new": row.price_toman,
                "method": method,
                "source": row.source,
            }
            for p, row, method, old in to_update[:15]
        ]
        return report

    with conn.cursor() as cur:
        for p, row, method, old in to_update:
            try:
                cur.execute(
                    "UPDATE products SET base_price = %s, updated_at = NOW() WHERE id = %s",
                    (row.price_toman, p["id"]),
                )
                report["updated"].append(
                    {
                        "id": p["id"],
                        "sku": p["sku"],
                        "old": str(old) if old is not None else None,
                        "new": row.price_toman,
                        "method": method,
                        "source": row.source,
                    }
                )
            except Exception as e:  # noqa: BLE001
                report["failed"].append({"sku": p["sku"], "error": str(e)})
        conn.commit()
    report["ok"] = len(report["updated"])
    report["fail"] = len(report["failed"])
    return report


def run_asal(conn, products, *, dry_run: bool) -> dict:
    rows = load_asal_excel(ASAL_XLSX)
    result = match_prices(products, rows, brand_id=INSIZE_BRAND_ID)
    applied = apply_updates(conn, result["matched"], dry_run=dry_run)
    return {
        "phase": "asal",
        "source": str(ASAL_XLSX),
        "list_codes": len(rows),
        "matched": len(result["matched"]),
        "unmatched_insize": len(result["unmatched_products"]),
        "unused_list_codes": len(result["unused_list_codes"]),
        "methods": dict(result["methods"]),
        "apply": applied,
        "unmatched_skus": [p["sku"] for p in result["unmatched_products"][:100]],
    }


def run_pdfs(conn, products, *, dry_run: bool) -> dict:
    per_file = []
    all_matched = []
    methods: Counter = Counter()
    total_codes = 0
    unused: list[PriceRow] = []

    for hint, brand_id, path in PDF_FILES:
        if not path.exists():
            per_file.append({"file": str(path), "error": "missing"})
            continue
        rows = parse_pdf_prices(path, hint)
        total_codes += len(rows)
        result = match_prices(products, rows, brand_id=brand_id)
        all_matched.extend(result["matched"])
        methods.update(result["methods"])
        unused.extend(result["unused_list_codes"])
        applied_preview = apply_updates(conn, result["matched"], dry_run=True)
        per_file.append(
            {
                "file": path.name,
                "brand_hint": hint,
                "brand_id": brand_id,
                "parsed_codes": len(rows),
                "priced": sum(1 for r in rows if r.price_toman is not None),
                "inquiry_zero": sum(1 for r in rows if r.price_toman is None),
                "matched": len(result["matched"]),
                "to_update": applied_preview["to_update"],
                "same": applied_preview["same"],
                "sample": [asdict(r) for r in rows[:5]],
                "update_sample": applied_preview.get("sample", [])[:5],
            }
        )

    # Deduplicate matched by product id — later files win
    by_pid: dict[int, tuple] = {}
    for p, row, method in all_matched:
        by_pid[p["id"]] = (p, row, method)
    deduped_matched = list(by_pid.values())
    applied = apply_updates(conn, deduped_matched, dry_run=dry_run)
    return {
        "phase": "pdf",
        "files": per_file,
        "list_codes": total_codes,
        "matched": len(deduped_matched),
        "unused_list_codes": len(unused),
        "methods": dict(methods),
        "apply": applied,
        "unused_sample": [asdict(r) for r in unused[:40]],
    }


def clear_placeholder_10m(conn, *, dry_run: bool) -> dict:
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(
            """
            SELECT id, sku, base_price FROM products
            WHERE deleted_at IS NULL AND base_price = 10000000
            """
        )
        rows = list(cur.fetchall())
        if dry_run:
            return {"count": len(rows), "sample_skus": [r["sku"] for r in rows[:20]]}
        cur.execute(
            """
            UPDATE products SET base_price = NULL, updated_at = NOW()
            WHERE deleted_at IS NULL AND base_price = 10000000
            """
        )
        conn.commit()
        return {"cleared": len(rows), "sample_skus": [r["sku"] for r in rows[:20]]}


def metrics(conn) -> dict:
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(
            """
            SELECT
              COUNT(*) FILTER (WHERE deleted_at IS NULL) AS alive,
              COUNT(*) FILTER (WHERE deleted_at IS NULL AND base_price IS NULL) AS no_price,
              COUNT(*) FILTER (WHERE deleted_at IS NULL AND base_price = 10000000) AS price_10m,
              COUNT(*) FILTER (
                WHERE deleted_at IS NULL
                  AND base_price IS NOT NULL
                  AND base_price > 0
                  AND base_price <> 10000000
              ) AS priced_real
            FROM products
            """
        )
        return dict(cur.fetchone())


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument(
        "--phase",
        choices=["all", "asal", "pdf", "clear10m", "metrics"],
        default="all",
    )
    args = ap.parse_args()
    if not args.dry_run and not args.apply and args.phase != "metrics":
        print("Specify --dry-run or --apply")
        return 2
    dry = args.dry_run or not args.apply

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    conn = db_connect()
    report: dict = {
        "at": datetime.now(timezone.utc).isoformat(),
        "dry_run": dry,
        "phase": args.phase,
        "metrics_before": metrics(conn),
    }
    products = load_products(conn)
    report["products_alive"] = len(products)

    if args.phase in ("all", "asal"):
        report["asal"] = run_asal(conn, products, dry_run=dry)
        products = load_products(conn)

    if args.phase in ("all", "pdf"):
        report["pdf"] = run_pdfs(conn, products, dry_run=dry)
        products = load_products(conn)

    if args.phase in ("all", "clear10m"):
        report["clear10m"] = clear_placeholder_10m(conn, dry_run=dry)

    report["metrics_after"] = metrics(conn)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    mode = "dry" if dry else "apply"
    out = REPORT_DIR / f"price_import_{args.phase}_{mode}_{stamp}.json"
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps({"wrote": str(out), "metrics_after": report["metrics_after"]}, ensure_ascii=False, indent=2))
    # compact phase summaries
    for key in ("asal", "pdf", "clear10m"):
        if key in report:
            block = report[key]
            if key == "clear10m":
                print(f"[{key}]", block)
            else:
                print(
                    f"[{key}] list={block.get('list_codes')} matched={block.get('matched')} "
                    f"apply={block.get('apply', {}).get('to_update') if dry else block.get('apply', {}).get('ok')}"
                )
    conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
