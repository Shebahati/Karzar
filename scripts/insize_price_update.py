#!/usr/bin/env python3
"""Update INSIZE product prices from Asal Toman price list Excel.

Matches Excel CODE to product SKU (exact, then base without trailing letters like A/S).
Uses TOMAN column as base_price.

Usage:
  /tmp/.xlsxvenv/bin/python scripts/insize_price_update.py --xlsx "/path/to/file.xlsx"
  /tmp/.xlsxvenv/bin/python scripts/insize_price_update.py --xlsx ... --dry-run
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path

API = os.getenv("KARZAR_API_BASE", "https://api.karzartools.com/api/v1")
INSIZE_BRAND_ID = 3
UA = "KarzarInsizePriceUpdater/1.0"


def _load_admin_creds() -> tuple[str, str]:
    phone = os.getenv("INITIAL_SUPER_ADMIN_PHONE")
    password = os.getenv("INITIAL_SUPER_ADMIN_PASSWORD")
    secrets = Path(__file__).resolve().parents[1] / ".deploy-secrets"
    if secrets.exists():
        for line in secrets.read_text(encoding="utf-8").splitlines():
            if line.startswith("INITIAL_SUPER_ADMIN_PHONE=") and not phone:
                phone = line.split("=", 1)[1].strip()
            if line.startswith("INITIAL_SUPER_ADMIN_PASSWORD=") and not password:
                password = line.split("=", 1)[1].strip()
    if not phone or not password:
        raise RuntimeError("Set INITIAL_SUPER_ADMIN_PHONE/PASSWORD or use .deploy-secrets")
    return phone, password


def http_json(method: str, url: str, *, data=None, headers=None, timeout=90):
    body = None
    hdrs = {"User-Agent": UA, "Accept": "application/json"}
    if headers:
        hdrs.update(headers)
    if data is not None:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        hdrs["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=body, headers=hdrs, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read()
            return resp.status, json.loads(raw.decode()) if raw else {}
    except urllib.error.HTTPError as e:
        raw = e.read().decode("utf-8", errors="ignore")
        try:
            payload = json.loads(raw) if raw else {}
        except json.JSONDecodeError:
            payload = {"raw": raw[:500]}
        return e.code, payload


def login() -> str:
    phone, password = _load_admin_creds()
    body = urllib.parse.urlencode({"username": phone, "password": password}).encode()
    req = urllib.request.Request(
        f"{API}/auth/login",
        data=body,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.loads(resp.read().decode())
    token = data.get("access_token")
    if not token:
        raise RuntimeError(f"login failed: {data}")
    return token


def norm(code: str) -> str:
    return str(code or "").strip().upper().replace(" ", "").replace("_", "-")


def code_variants(code: str) -> list[str]:
    c = norm(code)
    out = [c]
    m = re.match(r"^(\d{3,5}-\d+)([A-Z]+)$", c)
    if m:
        out.append(m.group(1))
    # 1205-2002S → 1205-2002
    m2 = re.match(r"^(\d{3,5}-\d+[0-9])([A-Z]+)$", c)
    if m2 and m2.group(1) not in out:
        out.append(m2.group(1))
    return list(dict.fromkeys(out))


def load_excel(path: Path) -> dict[str, dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: dict[str, dict] = {}
    for r in range(2, (ws.max_row or 0) + 1):
        code = ws.cell(r, 1).value
        desc = ws.cell(r, 2).value
        toman = ws.cell(r, 4).value
        if code is None or toman is None:
            continue
        try:
            price = int(float(toman))
        except (TypeError, ValueError):
            continue
        if price <= 0:
            continue
        key = norm(code)
        rows[key] = {"code": key, "price": price, "desc": desc, "row": r}
    return rows


def load_insize_products(auth: dict) -> list[dict]:
    products: list[dict] = []
    skip = 0
    limit = 1000
    while True:
        # Prefer admin list if available for id + base_price
        status, resp = http_json(
            "GET",
            f"{API}/products/?brand_id={INSIZE_BRAND_ID}&skip={skip}&limit={limit}",
            headers=auth,
        )
        if status != 200:
            raise RuntimeError(f"products list failed: {status} {resp}")
        items = resp.get("data") or []
        products.extend(items)
        total = (resp.get("meta") or {}).get("total_count")
        skip += len(items)
        if not items:
            break
        if total is not None and skip >= int(total):
            break
        if len(items) < limit:
            break
    return products


def match_products(products: list[dict], excel: dict[str, dict]):
    # index excel by base variant → list of codes
    by_base: dict[str, list[str]] = defaultdict(list)
    for code in excel:
        for v in code_variants(code):
            by_base[v].append(code)

    matched = []
    unmatched = []
    methods: Counter = Counter()

    for p in products:
        sku = norm(p.get("sku"))
        hit_code = None
        method = None
        if sku in excel:
            hit_code, method = sku, "exact"
        else:
            for v in code_variants(sku):
                if v in excel:
                    hit_code, method = v, "sku-as-excel-base"
                    break
                cands = by_base.get(v) or []
                # unique candidate for this base
                uniq = list(dict.fromkeys(cands))
                if len(uniq) == 1:
                    hit_code, method = uniq[0], "unique-base"
                    break
                if len(uniq) > 1:
                    # prefer candidate equal to sku, else shortest suffix delta
                    if sku in uniq:
                        hit_code, method = sku, "exact-among-cands"
                    else:
                        # prefer code that starts with sku + single letter suffix
                        letter = [c for c in uniq if c.startswith(sku) and len(c) == len(sku) + 1]
                        if len(letter) == 1:
                            hit_code, method = letter[0], "sku+letter"
                        else:
                            method = "ambiguous"
                    break
        if hit_code and method != "ambiguous":
            matched.append((p, excel[hit_code], method))
            methods[method] += 1
        else:
            unmatched.append(p)
            if method == "ambiguous":
                methods["ambiguous_skip"] += 1

    return matched, unmatched, methods


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--sleep", type=float, default=0.08)
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    excel = load_excel(xlsx)
    print(f"[price] excel codes={len(excel)} from {xlsx}")

    token = login()
    auth = {"Authorization": f"Bearer {token}"}
    products = load_insize_products(auth)
    print(f"[price] insize products={len(products)}")

    matched, unmatched, methods = match_products(products, excel)
    print(f"[price] matched={len(matched)} unmatched={len(unmatched)} methods={dict(methods)}")

    # prepare updates
    updates = []
    same = 0
    for p, row, method in matched:
        new_price = str(row["price"])
        old = p.get("base_price")
        try:
            old_f = float(old) if old is not None else None
        except (TypeError, ValueError):
            old_f = None
        if old_f is not None and abs(old_f - row["price"]) < 0.5:
            same += 1
            continue
        updates.append((p, row, method, old))

    print(f"[price] to_update={len(updates)} already_same={same}")
    if updates[:8]:
        print("[price] sample updates:")
        for p, row, method, old in updates[:8]:
            print(f"  {p.get('sku')} {old} → {row['price']} ({method})")

    report = {
        "excel_codes": len(excel),
        "insize_products": len(products),
        "matched": len(matched),
        "unmatched": len(unmatched),
        "already_same": same,
        "to_update": len(updates),
        "methods": dict(methods),
        "unmatched_skus": [p.get("sku") for p in unmatched],
        "updated": [],
        "failed": [],
    }

    if args.dry_run:
        Path("data/insize_price_update_report.json").write_text(
            json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print("[price] dry-run only → data/insize_price_update_report.json")
        return 0

    ok = fail = 0
    for i, (p, row, method, old) in enumerate(updates, 1):
        pid = p.get("id")
        payload = {"base_price": str(row["price"])}
        status, resp = http_json(
            "PUT", f"{API}/products/{pid}", data=payload, headers=auth
        )
        if status in (200, 201):
            ok += 1
            report["updated"].append(
                {"sku": p.get("sku"), "old": old, "new": row["price"], "method": method}
            )
            if ok % 25 == 0 or i == len(updates):
                print(f"[price] progress {i}/{len(updates)} ok={ok} fail={fail}")
        else:
            fail += 1
            msg = resp.get("message") or resp.get("error_code") or str(resp)[:200]
            report["failed"].append({"sku": p.get("sku"), "status": status, "msg": msg})
            if fail <= 10:
                print(f"[price] FAIL {p.get('sku')}: {status} {msg}")
        time.sleep(args.sleep)

    report["ok"] = ok
    report["fail"] = fail
    out = Path("data/insize_price_update_report.json")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[price] DONE ok={ok} fail={fail} → {out}")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
