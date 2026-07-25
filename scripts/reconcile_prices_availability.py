#!/usr/bin/env python3
"""Reconcile catalog prices from DOCS/Price and sync availability.

Rules (site unit = Toman):
  - Asal Excel TOMAN column → INSIZE base_price as-is
  - Markup CSVs `stored_price` → update matched SKUs (already Toman + markup)
  - Brand PDFs: Rial ÷ 10 → Toman (no extra markup unless CSV says so)
  - After price pass (or price-only skip): for ALL live products
      base_price present and > 0 → available
      else → unavailable
  - Schema adaptive:
      * if products.is_available exists → set boolean
      * else set stock_quantity to 1.0 / 0.0 (live staging semantics)

Usage:
  # Build payload only (no DB):
  .venv312/bin/python scripts/reconcile_prices_availability.py --build-payload

  # Apply via env DB (local / tunneled):
  .venv312/bin/python scripts/reconcile_prices_availability.py --apply

  # Dry-run against DB:
  .venv312/bin/python scripts/reconcile_prices_availability.py --dry-run

Never prints DB passwords. Prefer running inside API/db network on VPS.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

PRICE_DIR = Path(os.getenv("KARZAR_PRICE_DIR", "/home/moahmmad/Projects/Karzar/DOCS/Price"))
REPORT_DIR = Path(__file__).resolve().parents[1] / "data"
EXPORT_DATE = "2026-07-25"

INSIZE_BRAND_ID = 3
BRAND_HINTS: dict[str, int | None] = {
    "INSIZE": 3,
    "Asal": 3,
    "Dasqua": 4,
    "TERMA": 5,
    "ASTPOWER": 13,
    "AST-Power": 13,
    "SAN OU": 20,
    "San-Ou": 20,
    "TIGER TEC": 22,
    "Tiger-Tec": 22,
    "UTEX": 24,
    "Chumpower": 33,
    "3Keego": 35,
    "Keego": 35,
    "SHAMS": 47,
    "YOWAX": 48,
    "Mighty Seven": 49,
    "Mighty-Seven": 49,
    "ET": 50,
}

MIN_RIAL = 100_000
MIN_TOMAN = 10_000
# Use 10 so storefront shows «موجود» (not «موجودی محدود» which triggers below qty 10).
AVAILABLE_STOCK = Decimal("10")
UNAVAILABLE_STOCK = Decimal("0")

SKU_TOKEN_RE = re.compile(
    r"(?i)(?<![A-Z0-9/])("
    r"[A-Z]{1,6}-?[A-Z0-9]*/?[A-Z0-9][A-Z0-9._×x*]{1,40}"
    r"|\d{3,5}-\d{2,5}[A-Z]{0,3}"
    r"|\d{4,8}(?:/\d{1,3})?"
    r"|[A-Z]{2}\d{4,8}"
    r")(?![A-Z0-9-])"
)
PRICE_RE = re.compile(r"(?<![\d.])(\d{1,3}(?:,\d{3}){1,4}|\d{4,12})(?![\d.])")
JUNK_SKU = re.compile(
    r"(?i)^(mm|cm|kg|din|hsse|hss|iso|unf|unc|nc|nf|b\d{1,2}|m\d+(\.\d+)?|"
    r"ip\d+|vh\d+|hv\d+|tc\d+|page|row)$"
)


@dataclass
class PriceRow:
    code: str
    price_toman: int | None
    source: str
    brand_id: int | None = None
    raw_price: str = ""


def norm(code: str) -> str:
    c = str(code or "").strip().upper()
    c = c.replace(" ", "").replace("_", "-").replace("×", "X").replace("*", "X")
    c = c.replace("‐", "-").replace("–", "-").replace("—", "-")
    return c


def soft_key(code: str) -> str:
    return norm(code).replace("/", "-").replace(".", "")


def code_variants(code: str) -> list[str]:
    c = norm(code)
    out = [c]
    m = re.match(r"^(\d{3,5}-\d+)([A-Z]+)$", c)
    if m:
        out.append(m.group(1))
    if c.startswith("AST-"):
        out.append(c[4:])
    return list(dict.fromkeys(out))


def parse_price_number(raw: str) -> int | None:
    s = str(raw).strip().replace(",", "").replace("٬", "")
    if not s:
        return None
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


def rial_to_toman(rial: int) -> int:
    return int((Decimal(rial) / Decimal(10)).to_integral_value(rounding=ROUND_HALF_UP))


def db_connect():
    import psycopg2

    return psycopg2.connect(
        host=os.getenv("POSTGRES_SERVER", "127.0.0.1"),
        port=int(os.getenv("POSTGRES_PORT", "5435")),
        user=os.getenv("POSTGRES_USER", "postgres"),
        password=os.getenv("POSTGRES_PASSWORD", ""),
        dbname=os.getenv("POSTGRES_DB", "karzar_db"),
    )


def has_is_available(conn) -> bool:
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT 1 FROM information_schema.columns
            WHERE table_name = 'products' AND column_name = 'is_available'
            """
        )
        return cur.fetchone() is not None


def load_asal_excel(path: Path) -> list[PriceRow]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[PriceRow] = []
    for r in range(2, (ws.max_row or 0) + 1):
        code = ws.cell(r, 1).value
        toman = ws.cell(r, 4).value
        if code is None:
            continue
        key = norm(code)
        if not key:
            continue
        if toman is None:
            continue
        price = parse_price_number(toman)
        if price is None:
            continue
        rows.append(
            PriceRow(
                code=key,
                price_toman=None if price <= 0 else price,
                source=path.name,
                brand_id=INSIZE_BRAND_ID,
                raw_price=str(toman),
            )
        )
    by: dict[str, PriceRow] = {}
    for row in rows:
        by[row.code] = row
    return list(by.values())


def load_markup_csv(path: Path, brand_id: int | None) -> list[PriceRow]:
    rows: list[PriceRow] = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for raw in reader:
            sku = norm(raw.get("sku") or "")
            if not sku:
                continue
            stored = parse_price_number(raw.get("stored_price") or "")
            if stored is None or stored <= 0:
                continue
            if stored < MIN_TOMAN:
                continue
            brand_name = (raw.get("brand") or "").strip()
            bid = brand_id
            if brand_name.upper() in {"ET", "ای تی"}:
                bid = BRAND_HINTS["ET"]
            elif brand_name.upper() in {"ASTPOWER", "AST-POWER", "AST POWER"}:
                bid = BRAND_HINTS["ASTPOWER"]
            elif brand_name.upper() in {"CHUMPOWER"}:
                bid = BRAND_HINTS["Chumpower"]
            rows.append(
                PriceRow(
                    code=sku,
                    price_toman=stored,
                    source=path.name,
                    brand_id=bid,
                    raw_price=str(stored),
                )
            )
    by: dict[str, PriceRow] = {}
    for row in rows:
        by[row.code] = row
    return list(by.values())


def _looks_like_sku(token: str) -> bool:
    t = norm(token)
    if len(t) < 3 or len(t) > 48:
        return False
    if JUNK_SKU.match(t):
        return False
    if re.fullmatch(r"\d{1,3}", t):
        return False
    if re.fullmatch(r"\d{3,5}", t):
        return False
    if not re.search(r"\d", t):
        return False
    return True


def pdftotext(path: Path) -> str:
    proc = subprocess.run(
        ["pdftotext", "-layout", str(path), "-"],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    return proc.stdout


def parse_pdf_prices(path: Path, brand_id: int | None) -> list[PriceRow]:
    text = pdftotext(path)
    rows: list[PriceRow] = []
    for line in text.splitlines():
        prices = [(m.start(), m.group(1), parse_price_number(m.group(1))) for m in PRICE_RE.finditer(line)]
        money = []
        for start, raw, val in prices:
            if val is None:
                continue
            if "," in raw or val == 0 or val >= MIN_RIAL:
                money.append((start, raw, val))
        if not money:
            continue
        comma = [m for m in money if "," in m[1]]
        candidates = comma or [m for m in money if m[2] and m[2] >= MIN_RIAL]
        if not candidates:
            continue
        _, raw_price, rial = candidates[-1]
        if rial is None or (rial != 0 and rial < MIN_RIAL):
            continue
        left = line[: candidates[-1][0]]
        skus = [m.group(1) for m in SKU_TOKEN_RE.finditer(left) if _looks_like_sku(m.group(1))]
        if not skus:
            skus = [m.group(1) for m in SKU_TOKEN_RE.finditer(line) if _looks_like_sku(m.group(1))]
        if not skus:
            continue
        norms = [norm(s) for s in skus]

        def sku_score(t: str) -> tuple:
            return (
                1 if re.search(r"[A-Z]", t) and re.search(r"\d", t) else 0,
                1 if "-" in t or "/" in t else 0,
                len(t),
            )

        sku = max(norms, key=sku_score)
        toman = None if rial == 0 else rial_to_toman(rial)
        if toman is not None and toman < MIN_TOMAN:
            continue
        rows.append(
            PriceRow(
                code=sku,
                price_toman=toman,
                source=path.name,
                brand_id=brand_id,
                raw_price=raw_price,
            )
        )
    by: dict[str, PriceRow] = {}
    for row in rows:
        # Prefer higher price on duplicate SKU within one file
        prev = by.get(row.code)
        if prev is None:
            by[row.code] = row
        elif (row.price_toman or 0) > (prev.price_toman or 0):
            by[row.code] = row
    return list(by.values())


def discover_sources() -> list[tuple[str, int | None, Path, str]]:
    """Return (label, brand_id, path, kind) kind in {excel,csv,pdf}."""
    out: list[tuple[str, int | None, Path, str]] = []
    skip = {"_trash_2026-07-25", "_trash", "_incoming", "_misc"}

    asal = PRICE_DIR / "Asal" / f"{EXPORT_DATE}_Asal_Toman_v1405-04-27.xlsx"
    if asal.exists():
        out.append(("Asal/INSIZE", INSIZE_BRAND_ID, asal, "excel"))

    csv_specs = [
        ("AST-Power/endmills", BRAND_HINTS["ASTPOWER"], PRICE_DIR / "AST-Power" / f"{EXPORT_DATE}_AST-Power_endmills_markup20.csv"),
        ("AST-Power/holesaws", BRAND_HINTS["ASTPOWER"], PRICE_DIR / "AST-Power" / f"{EXPORT_DATE}_AST-Power_holesaws_markup10.csv"),
        ("AST-Power/tap-machines", BRAND_HINTS["ASTPOWER"], PRICE_DIR / "AST-Power" / f"{EXPORT_DATE}_AST-Power_tap-machines_markup10.csv"),
        ("ET/taps", BRAND_HINTS["ET"], PRICE_DIR / "ET" / f"{EXPORT_DATE}_ET_taps_markup10.csv"),
    ]
    for label, bid, path in csv_specs:
        if path.exists():
            out.append((label, bid, path, "csv"))

    # Brand PDFs (skip AST-Power PDFs already covered by azarsanat + markup CSVs
    # except we still parse non-AST folders that have reliable SKUs).
    pdf_folders = {
        "Dasqua": BRAND_HINTS["Dasqua"],
        "YOWAX": BRAND_HINTS["YOWAX"],
        "TERMA": BRAND_HINTS["TERMA"],
        "San-Ou": BRAND_HINTS["San-Ou"],
        "Mighty-Seven": BRAND_HINTS["Mighty-Seven"],
        "Tiger-Tec": BRAND_HINTS["Tiger-Tec"],
        "Keego": BRAND_HINTS["Keego"],
        "Chumpower": BRAND_HINTS["Chumpower"],
        "D-Coil": None,  # no brand row yet
        "UTEX": BRAND_HINTS["UTEX"],
        "ET": BRAND_HINTS["ET"],
        # AST-Power PDFs: optional extra pass; markup CSVs already authoritative for key families
        "AST-Power": BRAND_HINTS["ASTPOWER"],
    }
    for folder, bid in pdf_folders.items():
        d = PRICE_DIR / folder
        if not d.is_dir():
            continue
        for path in sorted(d.glob("*.pdf")):
            if any(p in path.parts for p in skip):
                continue
            # Skip scanned unknown with no extractable text priority
            if "Scanned-Unknown" in path.name:
                continue
            out.append((f"{folder}/{path.name}", bid, path, "pdf"))
    return out


def collect_price_rows() -> tuple[list[PriceRow], dict]:
    sources = discover_sources()
    all_rows: list[PriceRow] = []
    meta: dict = {"sources": [], "by_brand_folder": Counter()}
    for label, brand_id, path, kind in sources:
        try:
            if kind == "excel":
                rows = load_asal_excel(path)
            elif kind == "csv":
                rows = load_markup_csv(path, brand_id)
            else:
                rows = parse_pdf_prices(path, brand_id)
        except Exception as exc:  # noqa: BLE001
            meta["sources"].append(
                {"label": label, "path": str(path), "kind": kind, "error": str(exc)}
            )
            continue
        priced = sum(1 for r in rows if r.price_toman)
        meta["sources"].append(
            {
                "label": label,
                "path": str(path),
                "kind": kind,
                "rows": len(rows),
                "priced": priced,
                "brand_id": brand_id,
            }
        )
        meta["by_brand_folder"][label.split("/")[0]] += priced
        all_rows.extend(rows)

    # Merge: same (brand_id, code) → prefer higher priced; CSV/excel beat pdf on ties via order
    # Process in order: pdf first then csv/excel overwrite if >= 
    ranked = sorted(
        all_rows,
        key=lambda r: (
            0 if "markup" in r.source or r.source.endswith(".xlsx") or r.source.endswith(".csv") else 1,
            -(r.price_toman or 0),
        ),
    )
    merged: dict[tuple[int | None, str], PriceRow] = {}
    for row in ranked:
        key = (row.brand_id, row.code)
        if key not in merged:
            merged[key] = row
            continue
        prev = merged[key]
        # Prefer structured sources; else higher price
        prev_structured = prev.source.endswith((".csv", ".xlsx")) or "markup" in prev.source
        cur_structured = row.source.endswith((".csv", ".xlsx")) or "markup" in row.source
        if cur_structured and not prev_structured:
            merged[key] = row
        elif cur_structured == prev_structured and (row.price_toman or 0) > (prev.price_toman or 0):
            merged[key] = row
    meta["merged_codes"] = len(merged)
    meta["by_brand_folder"] = dict(meta["by_brand_folder"])
    return list(merged.values()), meta


def load_products(conn) -> list[dict]:
    from psycopg2.extras import RealDictCursor

    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(
            """
            SELECT p.id, p.sku, p.name, p.brand_id, p.base_price, p.stock_quantity,
                   b.name AS brand_name
            FROM products p
            LEFT JOIN brands b ON b.id = p.brand_id
            WHERE p.deleted_at IS NULL
            """
        )
        return list(cur.fetchall())


def match_rows(products: list[dict], price_rows: list[PriceRow]) -> dict:
    by_brand: dict[int | None, dict[str, PriceRow]] = defaultdict(dict)
    soft_by_brand: dict[int | None, dict[str, list[str]]] = defaultdict(lambda: defaultdict(list))
    for row in price_rows:
        by_brand[row.brand_id][row.code] = row
        soft_by_brand[row.brand_id][soft_key(row.code)].append(row.code)
        for v in code_variants(row.code):
            soft_by_brand[row.brand_id][soft_key(v)].append(row.code)

    matched = []
    methods: Counter = Counter()
    matched_pids: set[int] = set()
    used: set[tuple[int | None, str]] = set()

    def bind(p, row: PriceRow, method: str) -> bool:
        if p["id"] in matched_pids:
            return False
        matched.append((p, row, method))
        methods[method] += 1
        matched_pids.add(p["id"])
        used.add((row.brand_id, row.code))
        return True

    for p in products:
        sku = norm(p.get("sku"))
        bid = p.get("brand_id")
        # Prefer same-brand index, then unscoped (brand_id None), then any brand exact
        search_spaces: list[int | None] = []
        if bid is not None:
            search_spaces.append(bid)
        search_spaces.append(None)
        hit = None
        method = None
        for space in search_spaces:
            index = by_brand.get(space) or {}
            soft = soft_by_brand.get(space) or {}
            if sku in index:
                hit, method = index[sku], "exact"
                break
            soft_cands = list(dict.fromkeys(soft.get(soft_key(sku)) or []))
            if len(soft_cands) == 1 and soft_cands[0] in index:
                hit, method = index[soft_cands[0]], "soft-exact"
                break
            for v in code_variants(sku):
                if v in index:
                    hit, method = index[v], "variant"
                    break
            if hit:
                break
        if hit is None:
            # Cross-brand exact only when unique across all brand indexes
            cands = []
            for space, index in by_brand.items():
                if sku in index:
                    cands.append(index[sku])
            if len(cands) == 1:
                hit, method = cands[0], "cross-brand-unique"
        if hit is not None:
            bind(p, hit, method or "exact")

    unused = [r for r in price_rows if (r.brand_id, r.code) not in used]
    return {
        "matched": matched,
        "methods": methods,
        "unused_list_codes": unused,
        "matched_pids": matched_pids,
    }


def apply_price_updates(conn, matched, *, dry_run: bool) -> dict:
    to_update = []
    same = 0
    for p, row, method in matched:
        if row.price_toman is None:
            continue
        old = p.get("base_price")
        try:
            old_f = float(old) if old is not None else None
        except (TypeError, ValueError):
            old_f = None
        new = row.price_toman
        if old_f is not None and abs(old_f - new) < 0.5:
            same += 1
            continue
        to_update.append((p, row, method, old))

    report = {
        "same": same,
        "to_update": len(to_update),
        "updated": 0,
        "samples": [
            {
                "sku": p["sku"],
                "brand_id": p.get("brand_id"),
                "old": str(old) if old is not None else None,
                "new": row.price_toman,
                "method": method,
                "source": row.source,
            }
            for p, row, method, old in to_update[:25]
        ],
        "changes": [
            {
                "id": p["id"],
                "sku": p["sku"],
                "brand_id": p.get("brand_id"),
                "old": str(old) if old is not None else None,
                "new": row.price_toman,
                "method": method,
                "source": row.source,
            }
            for p, row, method, old in to_update
        ],
    }
    if dry_run or not to_update:
        return report

    with conn.cursor() as cur:
        for p, row, method, old in to_update:
            cur.execute(
                "UPDATE products SET base_price = %s, updated_at = NOW() WHERE id = %s AND deleted_at IS NULL",
                (row.price_toman, p["id"]),
            )
            report["updated"] += 1
        conn.commit()
    return report


def sync_availability(conn, *, dry_run: bool, use_bool: bool | None = None) -> dict:
    if use_bool is None:
        use_bool = has_is_available(conn)
    with conn.cursor() as cur:
        cur.execute(
            """
            SELECT
              COUNT(*) FILTER (WHERE base_price IS NOT NULL AND base_price > 0) AS priced,
              COUNT(*) FILTER (WHERE base_price IS NULL OR base_price = 0) AS unpriced,
              COUNT(*) AS alive
            FROM products WHERE deleted_at IS NULL
            """
        )
        priced, unpriced, alive = cur.fetchone()

        if use_bool:
            cur.execute(
                """
                SELECT
                  COUNT(*) FILTER (
                    WHERE base_price IS NOT NULL AND base_price > 0 AND is_available IS DISTINCT FROM TRUE
                  ),
                  COUNT(*) FILTER (
                    WHERE (base_price IS NULL OR base_price = 0) AND is_available IS DISTINCT FROM FALSE
                  )
                FROM products WHERE deleted_at IS NULL
                """
            )
            need_on, need_off = cur.fetchone()
            if not dry_run:
                cur.execute(
                    """
                    UPDATE products SET is_available = TRUE, updated_at = NOW()
                    WHERE deleted_at IS NULL AND base_price IS NOT NULL AND base_price > 0
                      AND is_available IS DISTINCT FROM TRUE
                    """
                )
                cur.execute(
                    """
                    UPDATE products SET is_available = FALSE, updated_at = NOW()
                    WHERE deleted_at IS NULL AND (base_price IS NULL OR base_price = 0)
                      AND is_available IS DISTINCT FROM FALSE
                    """
                )
                conn.commit()
                cur.execute(
                    """
                    SELECT
                      COUNT(*) FILTER (WHERE is_available IS TRUE),
                      COUNT(*) FILTER (WHERE is_available IS FALSE)
                    FROM products WHERE deleted_at IS NULL
                    """
                )
                avail, unavail = cur.fetchone()
            else:
                avail, unavail = priced, unpriced
            mode = "is_available"
        else:
            cur.execute(
                """
                SELECT
                  COUNT(*) FILTER (
                    WHERE base_price IS NOT NULL AND base_price > 0 AND stock_quantity <= 0
                  ),
                  COUNT(*) FILTER (
                    WHERE (base_price IS NULL OR base_price = 0) AND stock_quantity > 0
                  )
                FROM products WHERE deleted_at IS NULL
                """
            )
            need_on, need_off = cur.fetchone()
            if not dry_run:
                cur.execute(
                    """
                    UPDATE products SET stock_quantity = %s, updated_at = NOW()
                    WHERE deleted_at IS NULL AND base_price IS NOT NULL AND base_price > 0
                      AND stock_quantity <= 0
                    """,
                    (AVAILABLE_STOCK,),
                )
                cur.execute(
                    """
                    UPDATE products SET stock_quantity = %s, updated_at = NOW()
                    WHERE deleted_at IS NULL AND (base_price IS NULL OR base_price = 0)
                      AND stock_quantity <> 0
                    """,
                    (UNAVAILABLE_STOCK,),
                )
                conn.commit()
                cur.execute(
                    """
                    SELECT
                      COUNT(*) FILTER (WHERE stock_quantity > 0),
                      COUNT(*) FILTER (WHERE stock_quantity <= 0)
                    FROM products WHERE deleted_at IS NULL
                    """
                )
                avail, unavail = cur.fetchone()
            else:
                avail, unavail = priced, unpriced
            mode = "stock_quantity"

    return {
        "mode": mode,
        "alive": alive,
        "priced": priced,
        "unpriced": unpriced,
        "would_mark_available": need_on,
        "would_mark_unavailable": need_off,
        "available": avail,
        "unavailable": unavail,
        "dry_run": dry_run,
    }


def build_payload(path: Path) -> dict:
    rows, meta = collect_price_rows()
    payload = {
        "at": datetime.now(timezone.utc).isoformat(),
        "currency": "toman",
        "note": "PDF rial÷10; Asal toman as-is; markup CSV stored_price as-is",
        "meta": meta,
        "rows": [asdict(r) for r in rows if r.price_toman],
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def apply_from_payload(conn, payload: dict, *, dry_run: bool) -> dict:
    rows = [
        PriceRow(
            code=norm(r["code"]),
            price_toman=r.get("price_toman"),
            source=r.get("source") or "payload",
            brand_id=r.get("brand_id"),
            raw_price=str(r.get("raw_price") or ""),
        )
        for r in payload.get("rows") or []
        if r.get("price_toman")
    ]
    products = load_products(conn)
    result = match_rows(products, rows)
    price_report = apply_price_updates(conn, result["matched"], dry_run=dry_run)
    # Reload products for availability after price changes when applying
    if not dry_run:
        products = load_products(conn)
    avail = sync_availability(conn, dry_run=dry_run)
    unused = result["unused_list_codes"]
    by_brand_unused: Counter = Counter()
    for r in unused:
        by_brand_unused[str(r.brand_id)] += 1
    return {
        "list_priced_rows": len(rows),
        "matched": len(result["matched"]),
        "methods": dict(result["methods"]),
        "unused_list_skus": len(unused),
        "unused_by_brand_id": dict(by_brand_unused),
        "unused_sample": [asdict(r) for r in unused[:40]],
        "price": price_report,
        "availability": avail,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--build-payload", action="store_true")
    parser.add_argument("--payload", type=Path, default=REPORT_DIR / "price_reconcile_payload.json")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--availability-only", action="store_true")
    parser.add_argument("--report", type=Path, default=None)
    args = parser.parse_args()

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    if args.build_payload:
        payload = build_payload(args.payload)
        print(
            json.dumps(
                {
                    "payload": str(args.payload),
                    "merged_priced": len(payload["rows"]),
                    "sources": len(payload["meta"]["sources"]),
                    "by_folder": payload["meta"]["by_brand_folder"],
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 0

    if not args.apply and not args.dry_run and not args.availability_only:
        parser.error("Use --build-payload and/or --dry-run / --apply / --availability-only")

    conn = db_connect()
    try:
        if args.availability_only:
            report = {"availability": sync_availability(conn, dry_run=not args.apply)}
        else:
            if args.payload.exists() and not args.build_payload:
                payload = json.loads(args.payload.read_text(encoding="utf-8"))
            else:
                payload = build_payload(args.payload)
            report = apply_from_payload(conn, payload, dry_run=not args.apply)
            report["payload_meta"] = payload.get("meta")
        report["at"] = datetime.now(timezone.utc).isoformat()
        report["mode"] = "apply" if args.apply else "dry-run"
        out = args.report or (REPORT_DIR / f"price_reconcile_{'apply' if args.apply else 'dry'}_{ts}.json")
        out.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        summary = {
            "report": str(out),
            "mode": report["mode"],
            "matched": report.get("matched"),
            "price_updated": (report.get("price") or {}).get("updated"),
            "price_same": (report.get("price") or {}).get("same"),
            "price_to_update": (report.get("price") or {}).get("to_update"),
            "unused_list_skus": report.get("unused_list_skus"),
            "availability": report.get("availability"),
        }
        print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
