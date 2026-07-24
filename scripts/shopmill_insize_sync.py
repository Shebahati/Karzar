#!/usr/bin/env python3
"""Sync shopmilltools INSIZE crawl into Karzar.

- Exact shopmill product names
- Official INSIZE SKUs (already normalized in crawl)
- Update existing INSIZE rows (name + price when Asal list has code)
- Create missing products (brand=INSIZE, specs only, no marketing description/images)

Usage:
  python scripts/shopmill_insize_sync.py --jsonl data/shopmill_insize.jsonl --dry-run
  python scripts/shopmill_insize_sync.py --jsonl data/shopmill_insize.jsonl
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from pathlib import Path

API = os.getenv("KARZAR_API_BASE", "https://api.karzartools.com/api/v1")
INSIZE_BRAND_ID = 3
STOCK_QTY = "10"
UA = "KarzarShopmillInsizeSync/1.0"
ASAL_XLSX = Path("/home/moahmmad/KarZar/آصال تومان 27-04-1405.xlsx")

# More specific first — map shopmill category/name → Karzar selectable leaf
CATEGORY_RULES: list[tuple[tuple[str, ...], int]] = [
    (("سختی سنج", "hardness", "leeb", "rockwell", "vickers", "brinell"), 89),
    (("زبری", "roughness"), 93),
    (("ساعت شیطونک", "شیطانک", "dial-test", "test-indicator"), 60),
    (("ساعت اندیکاتور", "dial-indicator", "اندیکاتور"), 59),
    (("پایه ساعت", "dial-indicator-stand", "magnetic-stand"), 64),
    (("بور گیج", "سیلندر", "cylinder-gauge", "bore"), 63),
    (("عمق سنج", "depth-gauges", "depth-guage", "کولیس عمق"), 62),
    (("ارتفاع", "height"), 69),
    (("فیلر", "feeler"), 73),
    (("ضخامت سنج", "thickness"), 73),
    (("پین گیج", "pin-gauge"), 67),
    (("گیج رینگی", "ring-gauge", "go-no-go", "گیج توپی", "plug-gauge"), 67),
    (("راپورتر", "گیج بلوک", "gage-block", "gauge-block"), 66),
    (("شابلون", "pitch", "radius-gauge", "welding-gage", "taper"), 65),
    (("زاویه", "protractor"), 68),
    (("تراز", "level"), 71),
    (("گونیا", "square"), 75),
    (("پرگار", "compass", "caliper-gauge"), 72),
    (("خط کش", "steel-rule", "ruler"), 76),
    (("متر", "measuring-tape", "laser-distance"), 77),
    (("صفحه صافی", "surface-plate", "granite"), 79),
    (("میکرومتر", "micrometer"), 58),
    (("کولیس", "caliper"), 57),
    (("ست ابزار", "measurement-tool-kit", "tool-set"), 57),
    (("ترکمتر", "torque"), 61),
    (("قطعات یدکی", "accessories", "لوازم جانبی"), 80),
]


def _load_admin_creds() -> tuple[str, str]:
    phone = os.getenv("INITIAL_SUPER_ADMIN_PHONE")
    password = os.getenv("INITIAL_SUPER_ADMIN_PASSWORD")
    secrets = Path(__file__).resolve().parents[1] / ".deploy-secrets"
    if secrets.exists():
        for line in secrets.read_text(encoding="utf-8").splitlines():
            if line.startswith("INITIAL_SUPER_ADMIN_PHONE=") and not phone:
                phone = line.split("=", 1)[1].strip()
            if line.startswith("INITIAL_SUPER_ADMIN_PASSWORD=") and not password:
                password = line.split("=", 1)[1].strip()
    if not phone or not password:
        raise RuntimeError("missing admin creds")
    return phone, password


def http_json(method: str, url: str, *, data=None, headers=None, timeout=90):
    body = None
    hdrs = {"User-Agent": UA, "Accept": "application/json"}
    if headers:
        hdrs.update(headers)
    if data is not None:
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        hdrs["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=body, headers=hdrs, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read()
            return resp.status, json.loads(raw.decode()) if raw else {}
    except urllib.error.HTTPError as e:
        raw = e.read().decode("utf-8", errors="ignore")
        try:
            payload = json.loads(raw) if raw else {}
        except json.JSONDecodeError:
            payload = {"raw": raw[:500]}
        return e.code, payload


def login() -> str:
    phone, password = _load_admin_creds()
    body = urllib.parse.urlencode({"username": phone, "password": password}).encode()
    req = urllib.request.Request(
        f"{API}/auth/login",
        data=body,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.loads(resp.read().decode())
    return data["access_token"]


def map_category(row: dict) -> int:
    blob = " ".join(
        f"{c.get('slug') or ''} {c.get('name') or ''}"
        for c in (row.get("source_categories") or [])
    )
    blob = f"{blob} {row.get('name') or ''}".lower()
    for keys, cid in CATEGORY_RULES:
        if any(k.lower() in blob for k in keys):
            return cid
    return 57


def build_specifications(raw: dict) -> dict:
    tech = {
        "range": "",
        "accuracy": "",
        "resolution": "",
        "material": "",
        "standard": "",
        "battery_type": "",
    }
    extras: dict[str, str] = {}
    mapping = {
        "دامنه اندازه گیری": "range",
        "سایز": "range",
        "دقت اندازه گیری": "accuracy",
        "دقت": "accuracy",
        "تفکیک پذیری": "resolution",
        "تقسیم بندی اندازه": "resolution",
        "کشور سازنده": "material",
        "کد فنی": "standard",
    }
    for k, v in (raw or {}).items():
        key = str(k).strip()
        val = str(v).strip()
        if not val or key == "گارانتی":
            continue
        if key in mapping:
            tech[mapping[key]] = val
        else:
            extras[key] = val
    return {
        "technical_specs": tech,
        "features": {
            "waterproof": False,
            "data_output": False,
            "auto_power_off": False,
            "buttons": [],
            "certification": "",
        },
        "dimensions": {"L_mm": 0.0, "a_mm": 0.0, "b_mm": 0.0, "c_mm": 0.0, "d_mm": 0.0},
        "optional_accessories": [],
        "source_attributes": extras,
    }


def load_asal_prices() -> dict[str, int]:
    if not ASAL_XLSX.exists():
        return {}
    from openpyxl import load_workbook

    wb = load_workbook(ASAL_XLSX, data_only=True)
    ws = wb.active
    out: dict[str, int] = {}
    for r in range(2, (ws.max_row or 0) + 1):
        code = ws.cell(r, 1).value
        toman = ws.cell(r, 4).value
        if code is None or toman is None:
            continue
        try:
            price = int(float(toman))
        except (TypeError, ValueError):
            continue
        if price <= 0:
            continue
        out[str(code).strip().upper()] = price
    return out


def sku_variants(sku: str) -> list[str]:
    s = sku.upper()
    out = [s]
    m = re.match(r"^(\d+-\d+)([A-Z]+)$", s)
    if m:
        out.append(m.group(1))
    # also with common letter suffixes used in price lists
    if re.match(r"^\d+-\d+$", s):
        out.extend([s + "A", s + "S", s + "AC"])
    return list(dict.fromkeys(out))


def lookup_price(asal: dict[str, int], sku: str) -> int | None:
    for v in sku_variants(sku):
        if v in asal:
            return asal[v]
    # strip trailing letters from asal keys and compare
    base = re.sub(r"[A-Z]+$", "", sku.upper())
    for k, price in asal.items():
        if re.sub(r"[A-Z]+$", "", k) == base:
            return price
    return None


def load_insize_products(auth: dict) -> dict[str, dict]:
    """Map normalized sku → product (prefer brand INSIZE)."""
    by_sku: dict[str, dict] = {}
    skip = 0
    while True:
        st, resp = http_json(
            "GET",
            f"{API}/products/?brand_id={INSIZE_BRAND_ID}&skip={skip}&limit=1000",
            headers=auth,
        )
        if st != 200:
            raise RuntimeError(f"list insize failed {st} {resp}")
        items = resp.get("data") or []
        for p in items:
            sku = str(p.get("sku") or "").strip().upper()
            if sku:
                by_sku[sku] = p
        total = (resp.get("meta") or {}).get("total_count")
        skip += len(items)
        if not items:
            break
        if total is not None and skip >= int(total):
            break
        if len(items) < 1000:
            break
    return by_sku


def find_product(by_sku: dict[str, dict], sku: str) -> dict | None:
    for v in sku_variants(sku):
        if v in by_sku:
            return by_sku[v]
    return None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--jsonl", default="data/shopmill_insize.jsonl")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--sleep", type=float, default=0.1)
    ap.add_argument("--update-names", action="store_true", default=True)
    ap.add_argument("--no-update-names", action="store_true")
    args = ap.parse_args()
    update_names = not args.no_update_names

    rows = [
        json.loads(line)
        for line in Path(args.jsonl).read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    asal = load_asal_prices()
    print(f"[sync] shopmill rows={len(rows)} asal_prices={len(asal)}")

    token = login()
    auth = {"Authorization": f"Bearer {token}"}
    by_sku = load_insize_products(auth)
    print(f"[sync] existing insize={len(by_sku)}")

    to_create = []
    to_update_name = []
    to_update_price = []
    skip_no_sku = []
    already_ok = 0

    for row in rows:
        sku = (row.get("sku") or "").strip().upper()
        name = (row.get("name") or "").strip()[:255]
        if not sku:
            skip_no_sku.append(row)
            continue
        existing = find_product(by_sku, sku)
        price = lookup_price(asal, sku)

        if existing:
            need_name = update_names and (existing.get("name") or "") != name
            old_price = existing.get("base_price")
            try:
                old_f = float(old_price) if old_price is not None else None
            except (TypeError, ValueError):
                old_f = None
            need_price = price is not None and (
                old_f is None or abs(old_f - price) > 0.5
            )
            if need_name:
                to_update_name.append((existing, row, price))
            elif need_price:
                to_update_price.append((existing, row, price))
            else:
                already_ok += 1
        else:
            to_create.append((row, price))

    print(
        f"[sync] create={len(to_create)} update_name={len(to_update_name)} "
        f"update_price_only={len(to_update_price)} already_ok={already_ok} "
        f"no_sku={len(skip_no_sku)}"
    )
    if to_create[:8]:
        print("[sync] create sample:")
        for row, price in to_create[:8]:
            print(f"  {row['sku']} | {row['name'][:70]} | price={price}")
    if skip_no_sku:
        print("[sync] no-sku sample:")
        for row in skip_no_sku[:10]:
            print(f"  {row.get('name')[:80]} model={row.get('model_raw')}")

    if args.dry_run:
        Path("data/shopmill_insize_sync_plan.json").write_text(
            json.dumps(
                {
                    "create": len(to_create),
                    "update_name": len(to_update_name),
                    "update_price": len(to_update_price),
                    "already_ok": already_ok,
                    "no_sku": [r.get("name") for r in skip_no_sku],
                    "create_skus": [r["sku"] for r, _ in to_create],
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        print("[sync] dry-run → data/shopmill_insize_sync_plan.json")
        return 0

    ok = fail = 0
    failures: list[str] = []

    # updates (name ± price)
    for existing, row, price in to_update_name + [
        (e, r, p) for e, r, p in to_update_price
    ]:
        payload: dict = {}
        if (existing.get("name") or "") != row["name"][:255]:
            payload["name"] = row["name"][:255]
        if price is not None:
            payload["base_price"] = str(price)
        if not payload:
            continue
        st, resp = http_json(
            "PUT", f"{API}/products/{existing['id']}", data=payload, headers=auth
        )
        if st in (200, 201):
            ok += 1
        else:
            fail += 1
            failures.append(f"UPD {existing.get('sku')}: {st} {resp}")
        time.sleep(args.sleep)

    print(f"[sync] updates done ok={ok} fail={fail}")

    created = 0
    for i, (row, price) in enumerate(to_create, 1):
        cat = map_category(row)
        payload = {
            "sku": row["sku"][:50],
            "name": row["name"][:255],
            "category_id": cat,
            "brand_id": INSIZE_BRAND_ID,
            "base_price": str(price) if price is not None else None,
            "stock_quantity": STOCK_QTY,
            "stock_unit": "piece",
            "is_original": True,
            "is_active": True,
            "description": None,
            "pdf_catalog_url": None,
            "specifications": build_specifications(row.get("specifications") or {}),
        }
        st, resp = http_json("POST", f"{API}/products/", data=payload, headers=auth)
        if st in (200, 201):
            created += 1
            ok += 1
            if created % 40 == 0 or i == len(to_create):
                print(
                    f"[sync] create progress {i}/{len(to_create)} created={created} fail={fail}"
                )
        else:
            fail += 1
            msg = resp.get("message") or resp.get("error_code") or str(resp)[:200]
            failures.append(f"ADD {row['sku']}: {st} {msg}")
            if fail <= 12:
                print(f"[sync] FAIL {row['sku']}: {st} {msg}")
        time.sleep(args.sleep)

    report = {
        "created": created,
        "updated": ok - created,
        "fail": fail,
        "already_ok": already_ok,
        "no_sku": len(skip_no_sku),
        "failures": failures[:100],
    }
    Path("data/shopmill_insize_sync_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"[sync] DONE created={created} fail={fail} → data/shopmill_insize_sync_report.json")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
